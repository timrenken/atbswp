#! python3

## python3 multiplicationTable.py x 

import openpyxl, sys, logging
from openpyxl.styles import Font
from pathlib import Path

x = int(sys.argv[1])


table_file = Path.cwd() / "multiplicationTables.xlsx"

if table_file.exists():
    wb = openpyxl.load_workbook('multiplicationTables.xlsx')
else:
    wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = f"Multiplication Table for {x}"


for r in range(1,x+1):
    sheet.cell(row=1,column=r+1).value = r
    sheet.cell(row=1,column=r+1).font = Font(bold=True)
    sheet.cell(row=r+1,column=1).value = r
    sheet.cell(row=r+1,column=1).font = Font(bold=True)
    for c in range(1,x+1):
        sheet.cell(row=r+1,column=c+1).value = c*r


wb.save(table_file)